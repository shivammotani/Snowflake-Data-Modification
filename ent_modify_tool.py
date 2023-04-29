#Last Updated on 16/04/2023


########################################## Code Start ###############################################

########################################## Importing Libraries ###############################################

import tkinter as tk
from tkinter import filedialog
from openpyxl import  load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import pyodbc 
import os
import ctypes


#This code tries to improves the resolution of the tkinter window
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass


########################################## Functions ###############################################
#Fetching the current username 
def get_user_name():
    try:
        if os.name == 'nt':
            GetUserNameExW = ctypes.windll.secur32.GetUserNameExW
            name_display = 3
            size = ctypes.pointer(ctypes.c_ulong(0))
            GetUserNameExW(name_display, None, size)
            name_buffer = ctypes.create_unicode_buffer(size.contents.value)
            GetUserNameExW(name_display, name_buffer, size)
            return name_buffer.value
        else:
            import pwd
            # Note that for some reason pwd.getpwuid(os.geteuid())[4] did not work for me
            display_name = (entry[4] for entry in pwd.getpwall() if entry[2] == os.geteuid()).next()
            return display_name
    except:
        return "User"


#Function to try and test the ODBC connection.
def test_connection():
    try:
        global cnxn,cursor
        #Creating connection
        cnxn = pyodbc.connect(f'DSN=Snowflake_DSN;PWD={pwd.get()};Database=P_TEST;schema=public;warehouse=compute_wh')
        cursor = cnxn.cursor()
        cursor.fast_executemany = True
        #Exectuing SQL command using connection just created
        cursor.execute("Select current_version()")
        update_table_button_output.set("Connection Successful")
        update_table_button_display.config( fg= "green")
        read_excel_button['state'] = 'active'
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        if(sqlstate == '28000'):
            update_table_button_output.set("Invalid Username or Password")
            update_table_button_display.config( fg= "red")
            read_excel_button['state'] = 'disabled'
        else:
            update_table_button_output.set("Check ODBC connection parameters")
            update_table_button_display.config( fg= "red")
            read_excel_button['state'] = 'disabled'

#This code updates and shows the list of available sheet present within the workbook
def update_sheet():
    menu = dropdown_option["menu"]
    menu.delete(0, "end")
    for string in sheets:
        menu.add_command(label=string, 
            command=lambda value=string: excel_sheet_selected.set(value))

        
#  Function to select the Excel file
def select_excel_file():
    try:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])
        os.rename(file_path,file_path)
        selected_excel_file.set(file_path)
        selected_excel_file_a.set("Selected Excel file is: ")
        wb = load_workbook(file_path)
        sheets.clear()
        for sheet in wb.sheetnames:
            sheets.append(sheet)
        update_sheet()
        excel_sheet_selected_a.set("Selected Excel Sheet is: ")
        excel_file_label.config( fg= "black")
    except:
        selected_excel_file_a.set("Excel file is open. Please close the excel")
        excel_file_label.config( fg= "red")
        sheets.clear()
        update_sheet()
        selected_excel_file.set("")
        excel_file_validated.set("")
        excel_sheet_selected_a.set("")
        excel_sheet_selected.set("")
        update_table_button['state'] = 'disabled'

#Function to check if any duplicate entries are present in the excel
def checkForDuplicates(wbook, sheet):
    try:
        file_df = pd.read_excel(wbook,sheet)
        file_df = file_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        duplicate_rows = []
        # Columns on which pandas has to evaluate the duplicate data
        duplicate_row_index = file_df.duplicated(subset=["DATA_ORIGIN*","TGT_TBL_NM*","TGT_COL_NM*",
                                                    "UIR_TBL_NM_1","UIR_COL_NM_1","UIR_VAL_1","UIR_TBL_NM_2","UIR_COL_NM_2","UIR_VAL_2","UIR_TBL_NM_3","UIR_COL_NM_3","UIR_VAL_3",
                                                    "UIR_TBL_NM_4","UIR_COL_NM_4","UIR_VAL_4","UIR_TBL_NM_5","UIR_COL_NM_5","UIR_VAL_5","UIR_TBL_NM_6","UIR_COL_NM_6","UIR_VAL_6",
                                                    "UIR_TBL_NM_7","UIR_COL_NM_7","UIR_VAL_7","UIR_TBL_NM_8","UIR_COL_NM_8","UIR_VAL_8","UIR_TBL_NM_9","UIR_COL_NM_9","UIR_VAL_9"],
                                                    keep="first")
        for i in range(len(duplicate_row_index)):
            if(duplicate_row_index[i]):
                duplicate_rows.append(i+2)
        # Returning the index where pandas found duplicates. This follows 0 based indexing skipping header(1st row)
        return duplicate_rows
    except:
        pass
    
    
# Function to evaluate the selected excel sheet before modifying anything via the ODBC
def validate_sheet(*args):
    try:
        excel_file_validated.set("")
        update_table_button_percent.set("")
        update_table_button_output.set('')
        update_table_button_error.set('')
        wbook = load_workbook(selected_excel_file.get())
        sheet_choosed = wbook[excel_sheet_selected.get()]
        rows = sheet_choosed.max_row
        noOfDuplicates = []
        if(sheet_choosed.max_column == 32):
            noOfDuplicates = checkForDuplicates(selected_excel_file.get(),excel_sheet_selected.get())

        # Checking if the selected excel doesn't have all the important columns and whether it's in correct format
        if(sheet_choosed.max_column != 32):
            update_table_button['state'] = 'disabled'
            excel_file_validated.set("Invalid template. Please use the valid template")
            excel_file_label_b.config( fg= "red")
            return
        
        # Checking if the selected excel has any duplicates
        elif(len(noOfDuplicates)):
            update_table_button['state'] = 'disabled'
            msg = "{dupl} Duplicates present in the sheet. \nAll duplicates have been highlighted in yellow".format(dupl = len(noOfDuplicates))
            excel_file_validated.set(msg)
            excel_file_label_b.config( fg= "red")
            for i in noOfDuplicates:
                sheet_choosed['A'+str(i)].fill = PatternFill("solid", start_color="FFFFFF00")
            wbook.save(selected_excel_file.get())
            return

        else:
            for row in range(2,rows+1):
                for col in range(1,5):
                    char = get_column_letter(col)
                    # Checking if cells are empty or contains NULL for key columns
                    if(sheet_choosed[char + str(row)].value == None or sheet_choosed[char + str(row)].value == ' ' or sheet_choosed[char + str(row)].value == '' or
                    str(sheet_choosed[char + str(row)].value).upper() == 'NULL'):
                        update_table_button['state'] = 'disabled'
                        update_table_button_output.set("Enter Value for Key Columns")
                        update_table_button_display.config(fg = 'red')
                        return
            update_table_button['state'] = 'active'
            excel_file_validated.set("Valid template")
            update_table_button.config( fg= "green")
            excel_file_label_b.config( fg= "green")
            update_table_button_display.config( fg= "green")
            update_frame.update()
          
    except:
        pass


# Function to display the entered password
def show_pwd():
    if(Checkbutton1.get()==1):
        test_connection_pwd.config(show='') 
    else:
        test_connection_pwd.config(show='*') 

# Function to check whether the record already exist in the database or not. This will be called both before inserting and deleting records.
def check_data_before_modifying(env, value_array):
    final_query = """Select count(*) from "{env}"."PUBLIC"."ENT_LKUP_DATA" where """.format(env = env)
    for i in range(0,32):
        if(i == 3 or i == 4):
            continue
        elif(i == 31):
            final_query = final_query + column_headers[i] + " = " + "'" + value_array[i] + "'" + " ; "
        else:
            final_query = final_query + column_headers[i] + " = " + "'" + value_array[i] + "'" + " and "
    try:
        row = cursor.execute(final_query)
        row = row.fetchone()[0]
        if(row == 0):
            return False
        else:
            return True  
    except:
        pass


# Function to insert records in the database using the excel selected
def insert_into_table(env): 
    try:
        counter = 0
        skipped = 0
        values = []
        wbook = load_workbook(selected_excel_file.get())
        sheet_choosed = wbook[excel_sheet_selected.get()]
        rows = sheet_choosed.max_row
        update_table_button['state'] = 'disabled'

        # Creating the SQL statement
        sql = """Insert into "{env}"."PUBLIC"."ENT_LKUP_DATA" ( """.format(env = env)
        for i in range(0,32):
            if(i != 31):
                sql = sql +  column_headers[i] + ", "
            else:
                sql = sql + column_headers[i] + " )\n values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"

        # Iterating through all the rows present in the excel
        for row in range(2,rows+1):
            val = []
            for col in range(1,33):
                char = get_column_letter(col)
                if(sheet_choosed[char + str(row)].value == None or sheet_choosed[char + str(row)].value == ' ' or sheet_choosed[char + str(row)].value == '' or
                   str(sheet_choosed[char + str(row)].value).upper() == 'NULL'):
                    val.append("NULL")
                else:
                    st = str(sheet_choosed[char + str(row)].value).strip()
                    val.append(st)
            try:
                # If value is already present in databse  then don't insert
                if(check_data_before_modifying(env,val) == True):
                    skipped = skipped + 1
                    warn = "Skipped {skipped} row. Record already exist in table".format(skipped = skipped)
                    update_table_button_error_display.config(fg= "red")
                    update_table_button_error.set(warn)
                    update_frame.update()
                    sheet_choosed['A'+str(row)].fill = PatternFill("solid", start_color="FFFF0000")

                else:
                    # Adding all valid rows to be inserted in master list
                    values.append(val)
                    counter = counter + 1
                    txt = "Inserting {counter} rows into {env}.ENT_LKUP_DATA".format(counter = counter, env = env)
                    update_table_button_output.set(txt)
                    update_frame.update()
                    sheet_choosed['A'+str(row)].fill = PatternFill(fill_type=None,start_color="FFFFFF")
                    
                total = skipped + counter
                percent_complete = (float)((total/(rows-1))*100)
                percent_complete = float("{:.2f}".format(percent_complete))
                if(total != rows-1):
                    msg = "Updating ENT_LKUP_DATA...{per}% done".format(per = percent_complete)
                    update_table_button_percent.set(msg)
                    update_table_button_percent_display.config( fg= "blue")
                    update_frame.update()

            except:
                cnxn.rollback()
                wbook.save(selected_excel_file.get())
                update_table_button_output.set("Error Occured. Validate your Data")
                return
        try:
            if(total == rows-1):
                # Inserting all the valid values from the master list to the database               
                for i in range (0,len(values),25):
                    cursor.executemany(sql,values[i:i+25])
                cnxn.commit()
                txt = "Inserted {counter} rows into {env}.ENT_LKUP_DATA".format(counter = counter, env = env)
                update_table_button_output.set(txt)
                update_frame.update()
                msg = "Update Complete...100% done"
                update_table_button_percent.set(msg)
                update_table_button_percent_display.config( fg= "green")
                update_frame.update()
            
        except:
            msg = "Error Occured. Update Rolled Back"
            update_table_button_percent.set(msg)
            update_table_button_percent_display.config( fg= "red")
            update_frame.update()
            cnxn.rollback()
        wbook.save(selected_excel_file.get())

    except:
        pass 
    
# Function to delte records in the database using the excel selected
def delete_from_table(env):
    try:
        file_name = os.path.basename(selected_excel_file.get())
        skipped = 0
        counter = 0
        values = []
        wbook = load_workbook(selected_excel_file.get())
        sheet_choosed = wbook[excel_sheet_selected.get()]
        rows = sheet_choosed.max_row
        update_table_button['state'] = 'disabled'

        # Creating the SQL statement
        sql = """Delete from "{env}"."PUBLIC"."ENT_LKUP_DATA" where
                DATA_ORIGIN=  ? and   TGT_TBL_NM=  ? and   TGT_COL_NM=  ? and   TGT_VAL=  ? and   TGT_DESC=  ? and   UIR_TBL_NM_1=  ? and   UIR_COL_NM_1=  ? and   UIR_VAL_1=  ? and   
                UIR_TBL_NM_2=  ? and   UIR_COL_NM_2=  ? and   UIR_VAL_2=  ? and   UIR_TBL_NM_3=  ? and   UIR_COL_NM_3=  ? and   UIR_VAL_3=  ? and   UIR_TBL_NM_4=  ? and   
                UIR_COL_NM_4=  ? and   UIR_VAL_4=  ? and   UIR_TBL_NM_5=  ? and   UIR_COL_NM_5=  ? and   UIR_VAL_5=  ? and   UIR_TBL_NM_6=  ? and   UIR_COL_NM_6=  ? and   UIR_VAL_6=  ?
                and   UIR_TBL_NM_7=  ? and   UIR_COL_NM_7=  ? and   UIR_VAL_7=  ? and   UIR_TBL_NM_8=  ? and   UIR_COL_NM_8=  ? and   UIR_VAL_8=  ? and  
                UIR_TBL_NM_9=  ? and   UIR_COL_NM_9=  ? and   UIR_VAL_9 =  ? """.format(env = env)
        
        # Iterating through all the rows present in the excel
        for row in range(2,rows+1):
            val = []
            for col in range(1,33):
                char = get_column_letter(col)
                if(sheet_choosed[char + str(row)].value == None or sheet_choosed[char + str(row)].value == ' ' or sheet_choosed[char + str(row)].value == '' or
                   str(sheet_choosed[char + str(row)].value).upper() == 'NULL'):
                    val.append("NULL")
                else:
                    st = str(sheet_choosed[char + str(row)].value).strip()
                    val.append(st) 

            try:
                # If value is not present in databse then don't do delete
                if(check_data_before_modifying(env,val) == False):
                    skipped = skipped + 1
                    warn = "Skipped {skipped} row. Record does not exist in table".format(skipped = skipped)
                    update_table_button_error_display.config(fg= "red")
                    update_table_button_error.set(warn)
                    update_frame.update()
                    sheet_choosed['A'+str(row)].fill = PatternFill("solid", start_color="FFFF0000")
                
                else:
                    # Adding all valid rows to be inserted in master list
                    values.append(val)
                    counter = counter + 1
                    txt = "Deleting {counter} row from {env}.ENT_LKUP_DATA.....".format(counter = counter, env = env)
                    update_table_button_output.set(txt)
                    sheet_choosed['A'+str(row)].fill = PatternFill(fill_type=None,start_color="FFFFFF")
                    update_frame.update()

                total = skipped + counter
                percent_complete = (float)((total/(rows-1))*100)
                percent_complete = float("{:.2f}".format(percent_complete))
                if(total != rows-1):
                    msg = "Updating ENT_LKUP_DATA...{per}% done".format(per = percent_complete)
                    update_table_button_percent.set(msg)
                    update_table_button_percent_display.config( fg= "blue")
                    update_frame.update()

            except:
                cnxn.rollback()
                wbook.save(file_name)
                update_table_button_output.set("Error Occured. Validate your Data")
                return
        try:
            if(total == rows-1):
                # Deleting all the valid values from the database
                for i in range(0,len(values)):
                    cursor.execute(sql,values[i])
                cnxn.commit()
                txt = "Deleted {counter} rows from {env}.ENT_LKUP_DATA.....".format(counter = counter, env = env)
                update_table_button_output.set(txt)
                update_frame.update()
                msg = "Update Complete...100% done"
                update_table_button_percent.set(msg)
                update_table_button_percent_display.config( fg= "green")
                update_frame.update()      
            
        except:
            msg = "Error Occured. Update Rolled Back"
            update_table_button_percent.set(msg)
            update_table_button_percent_display.config( fg= "red")
            update_frame.update()
            cnxn.rollback()

        wbook.save(selected_excel_file.get())
    except:
        pass 

    
    
# Function to select schema and call appropriate function based on user choice
def update_table():
    env = ''
    if(env_select_radio.get() == 'DEV'):
        env = 'P_DEV'
    elif(env_select_radio.get() == 'TEST'):
        env = 'P_TEST'
    elif(env_select_radio.get() == 'CERT'):
        env = 'P_CERT'
    else:
        env = 'P_PROD'

    if(data_mod_radio.get() == 'INSERT' == 'INSERT'):
        insert_into_table(env)
    elif(data_mod_radio.get() == 'DELETE'):
        delete_from_table(env)


########################################## Functions END ###############################################


########################################## Tkinter Initializtion ###############################################

# Create the tkinter window
window = tk.Tk()
window.resizable(False,False)
window.geometry('700x450')
username = get_user_name()
if(username.find("[") != -1):
   username = username[:username.find("[")]
window.title("""Welcome to ENT Data Modification Tool - {username}""".format(username = username))


########################################## Tkinter Initializtion End ###############################################


########################################## Global Variables ###############################################


cnxn = ''
cursor = ''
selected_excel_file = tk.StringVar()
selected_excel_file_a = tk.StringVar()
excel_sheet_selected = tk.StringVar()
excel_sheet_selected_a = tk.StringVar()
excel_file_validated = tk.StringVar()
update_table_button_output = tk.StringVar()
update_table_button_error = tk.StringVar()
update_table_button_percent = tk.StringVar()
pwd = tk.StringVar()
Checkbutton1 = tk.IntVar(value=0)
excel_sheet_selected.trace('w',validate_sheet)
data_mod_radio = tk.StringVar(None, "INSERT")
env_select_radio = tk.StringVar(None, "DEV")
sheets = ['']
sql_commands = []
column_headers = ["DATA_ORIGIN","TGT_TBL_NM", "TGT_COL_NM", 
                        "TGT_VAL", "TGT_DESC", 
                        "UIR_TBL_NM_1", "UIR_COL_NM_1", "UIR_VAL_1",
                        "UIR_TBL_NM_2", "UIR_COL_NM_2", "UIR_VAL_2", 
                        "UIR_TBL_NM_3", "UIR_COL_NM_3", "UIR_VAL_3",
                        "UIR_TBL_NM_4", "UIR_COL_NM_4", "UIR_VAL_4",
                        "UIR_TBL_NM_5", "UIR_COL_NM_5", "UIR_VAL_5",
                        "UIR_TBL_NM_6", "UIR_COL_NM_6", "UIR_VAL_6",
                        "UIR_TBL_NM_7", "UIR_COL_NM_7", "UIR_VAL_7",
                        "UIR_TBL_NM_8", "UIR_COL_NM_8", "UIR_VAL_8",
                        "UIR_TBL_NM_9", "UIR_COL_NM_9", "UIR_VAL_9"]


########################################## Global Variables End ###############################################


########################################## Tkinter Objects ###############################################


# Button to test connection and select the Excel file and read it
frame_a = tk.Frame(window)
frame_a.grid(row=0,column=0, sticky="EW")
test_connection_label = tk.Label(frame_a, text="Enter your password: ")
test_connection_label.grid(row=2, column=0,sticky="EW")
test_connection_pwd = tk.Entry(frame_a, show="*", width=20, textvariable=pwd)
test_connection_pwd.grid(row=2, column=1)
test_connection_checkbox = tk.Checkbutton(frame_a, text = "Show Password", 
                      variable = Checkbutton1, onvalue = 1,offvalue = 0,command=show_pwd)
test_connection_checkbox.grid(row=2, column=2,sticky="EW")
test_connection_button = tk.Button(frame_a, text="Test Connection", command=test_connection)
test_connection_button.grid(row=2, column=3)
read_excel_button = tk.Button(frame_a, text="Read Excel file", command=select_excel_file)
read_excel_button.grid(row=3, column=0,sticky="EW")
# read_excel_button['state'] = 'disabled'
excel_file_label = tk.Label(frame_a, textvariable=selected_excel_file_a)
excel_file_label.grid(row=3, column=1,sticky="EW")
excel_file_label_a = tk.Label(frame_a, textvariable=selected_excel_file)
excel_file_label_a.grid(row=3, column=2,sticky="EW")

for child in frame_a.winfo_children():
    child.grid_configure(padx=10, pady=10)




#Dropdown to select the sheet within the Excel file
frame_dropdown = tk.Frame(window)
frame_dropdown.grid(row=1,column=0,sticky="EW")
dropdown_option = tk.OptionMenu( frame_dropdown , excel_sheet_selected , *sheets)
dropdown_option.grid(row=0, column=0)
sheet_file_label = tk.Label(frame_dropdown, textvariable=excel_sheet_selected_a)
sheet_file_label.grid(row=0, column=1,sticky="EW")
sheet_file_label_a = tk.Label(frame_dropdown, textvariable=excel_sheet_selected)
sheet_file_label_a.grid(row=0, column=2,sticky="EW")
excel_file_label_b = tk.Label(frame_dropdown, textvariable=excel_file_validated)
excel_file_label_b.grid(row=0, column=3,sticky="EW")
for child in frame_dropdown.winfo_children():
    child.grid_configure(padx=10, pady=5)


# Label for radio buttons
data_mod_radio_frame = tk.Frame(window)
data_mod_radio_frame.grid(row=2,column=0,sticky="EW",pady=10)
radio_button_label = tk.Label(data_mod_radio_frame, text="Select an option:")
radio_button_label.grid(row=0, column=0)

# Radio buttons for Data Modification 
#Insert
radio_button_insert = tk.Radiobutton(data_mod_radio_frame, text="Insert", variable=data_mod_radio, value="INSERT")
radio_button_insert.grid(row=1, column=0)

#Delete
radio_button_delete = tk.Radiobutton(data_mod_radio_frame, text="Delete", variable=data_mod_radio, value="DELETE")
radio_button_delete.grid(row=1, column=1)

for child in data_mod_radio_frame.winfo_children():
    child.grid_configure(padx=10, pady=3,sticky="W")


# Radio buttons for ENV Selection
env_radio_frame = tk.Frame(window)
env_radio_frame.grid(row=3,column=0,sticky="EW",pady=5)

#DEV
radio_button_dev = tk.Radiobutton(env_radio_frame, text="Dev", variable=env_select_radio, value="DEV")
radio_button_dev.grid(row=0, column=0)

#TEST
radio_button_test = tk.Radiobutton(env_radio_frame, text="Test", variable=env_select_radio, value="TEST")
radio_button_test.grid(row=0, column=1)

#CERT
radio_button_cert = tk.Radiobutton(env_radio_frame, text="Cert", variable=env_select_radio, value="CERT")
radio_button_cert.grid(row=0, column=2)

#PROD
radio_button_prod = tk.Radiobutton(env_radio_frame, text="Prod", variable=env_select_radio, value="PROD")
radio_button_prod.grid(row=0, column=3)

for child in env_radio_frame.winfo_children():
    child.grid_configure(padx=10, pady=5,sticky="EW")


# Button to start pushing the data into the table based on radio button selection
update_frame = tk.Frame(window)
update_frame.grid(row=4,column = 0,pady=10)
update_table_button = tk.Button(update_frame, text="Update table", command=update_table)
update_table_button.grid(row = 0, column=0, padx=10, pady=5,sticky="EW")
update_table_button_display = tk.Label(update_frame, textvariable=update_table_button_output)
update_table_button_display.grid(row=2, column=0,sticky="EW")
update_table_button_error_display = tk.Label(update_frame, textvariable=update_table_button_error)
update_table_button_error_display.grid(row=3, column=0,sticky="EW")
update_table_button_percent_display = tk.Label(update_frame, textvariable=update_table_button_percent)
update_table_button_percent_display.grid(row=4, column=0,sticky="EW")
update_table_button['state'] = 'disabled'


# Start the tkinter event loop
window.mainloop()
cursor.close()
cnxn.close()

########################################## Tkinter Objects ###############################################

########################################## Code End ###############################################






